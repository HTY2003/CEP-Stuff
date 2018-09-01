from openpyxl import *
import random

student_number = 0
choice_number = 0

wb = load_workbook("cleaned.xlsx")
classlist = wb["classlist"]
choiceslist = wb["choices"]
MEP = wb["MEP"]
DSA = wb["DSA"]
# The psychomotor spreadsheet will not be used because a student's desire to join a CCA
# And his athletic ability can be determined during the CCA trials

musiclist = ["SE","RV","GE","RIMB","RICO","RP"]
# Used for the allocate_MEP function

ccaswithshortlist = list(wb.get_sheet_names())
# Used to check if the cca has a shortlist for adding students
# E.g swimming has no shortlist and hence no spreadsheet
if "CO" in ccaswithshortlist:
    ccaswithshortlist.remove("CO")
    ccaswithshortlist.append("RICO")

slassignlist = []

studentlist=[]
for row in classlist.iter_rows(min_row=2,min_col=3,max_row=500,max_col=3):
    for cell in row:
        if cell.value != None:
            studentlist.append(cell.value)
            student_number = student_number + 1
for row in choiceslist.iter_rows(min_row=2,min_col=2,max_row=500,max_col=2):
    for cell in row:
        if cell.value != None:
            choice_number = choice_number + 1

namechangedict = {}
# Made to account for the inconsistences in name between spreadsheets
namechangedict["01SCOUT"] = "O1"
namechangedict["02SCOUT"] = "O2"

quotadict = {}
# Keeps track of all current CCA Quotas
quota=wb["ccaquota"]
for i in range(1,50):
    if isinstance(quota["E"+str(i)].value, int):
        if quota["B"+str(i)].value == "RD":
                quotadict["Debater"] = quota["E"+str(i)].value
        else:
            quotadict[quota["B"+str(i)].value] = quota["E"+str(i)].value

#-------------------------FUNCTIONS-----------------------

def allocate(name, cca, choice, rank = None):
    if name in studentlist:
        studentlist.remove(name)
# First, the student info is updated in the classlist spreadsheet
        for i in range(2,2+student_number):
            if classlist["C"+str(i)].value == name:
                classlist["G"+str(i)] = cca
                classlist["H"+str(i)] = choice
                classlist["I"+str(i)] = rank
# Next, everyone who got allocated by shortlist is recorded for the stat counter later.
        if rank != None:
            slassignlist.append(name)
            if cca in ccaswithshortlist:
                if cca == "RICO":
                    shortlist = wb["CO"]
                else:
                    shortlist = wb[cca]
# Finally, the quota stats on the spreadsheet and dictionary are updated
        quotadict[cca] = quotadict[cca] - 1
        for q in range(1,50):
            if cca == "Debater":
                if quota["B"+str(q)].value == "RD":
                    quota["D"+str(q)] = quota["D"+str(q)].value + 1
            elif quota["B"+str(q)].value == cca:
                quota["D"+str(q)] = quota["D"+str(q)].value + 1
        print(name)


def allocate_dsa():
# This function allocates DSA students to their respective CCAs.
    print("DSA")
    for i in range(2,80):
        name = DSA["A"+str(i)].value
        cca = DSA["C"+str(i)].value
        allocate(name, cca, "DSA")

def allocate_mep():
# This function allocates MEP students based on whether 1)their first choice is music-based
# And 2)if they are in the shortlist for the CCA.
    print("MEP")
    for i in range(2,50):
        name = MEP["A"+str(i)].value
        for n in range(2,2+choice_number):
            if name == choiceslist["B"+str(n)].value:
                cca = choiceslist["F"+str(n)].value
                if cca in musiclist:
                    if cca == "RICO":
                        shortlist = wb["CO"]
                    else:
                        shortlist = wb[cca]
                    for r in range(2,40):
                        if shortlist["B"+str(r)].value == name:
                            if quotadict[cca] > 0:
                                allocate(name, cca, 1, rank = r-1)

def allocate_restshortlist():
# This function scans through the top rank in every CCA and if the person has put it as their first choice, then it allocates.
# Then the same statement is carried out except for each second ranked student.
# When every rank has been scanned, everything repeats but with the second choice.
    print("Rest:Shortlist")
    for c in range(6,9):
        for r in range(2,40):
            for cca in list(quotadict.keys()):
                if cca in ccaswithshortlist:
                    for n in range(2,2+choice_number):
                        if cca == "RICO": shortlist=wb["CO"]
                        else: shortlist=wb[cca]
                        name = shortlist["B"+str(r)].value
                        if name == choiceslist["B"+str(n)].value:
                            if choiceslist.cell(row=n,column=c).value in list(namechangedict.keys()):
                                choice = namechangedict[choiceslist.cell(row=n,column=c).value]
                            else:
                                choice = choiceslist.cell(row=n,column=c).value
                            if choice == cca:
                                if quotadict[cca] > 0:
                                    allocate(name, cca, c-5, rank = r-1)

def allocate_restnormal():
# If slots are open in a first choice, they get allocated. If not, next choice.
    print("Rest:Normal")
    for c in range(6,15):
        for s in studentlist:
            for n in range(2,2+choice_number):
                if s == choiceslist["B"+str(n)].value:
                    cca = choiceslist.cell(row=n,column=c).value
                    if cca in list(namechangedict.keys()):
                        if quotadict[namechangedict[cca]] > 0:
                            allocate(s, namechangedict[cca], c-5)
                    elif quotadict[cca] > 0:
                            allocate(s, cca, c-5)

def allocate_restrandom():
# This  function is for those who did not hand in a choice or got really unlucky and all their choices' quotas were filled.
    print("Rest:Random")
# First, all remaining CCA slots are allocated.
    for cca in list(quotadict.keys()):
        if quotadict[cca] > 0:
            for repeat in range(quotadict[cca]):
                allocate(studentlist[0], cca, None)
# Next, a slot is added to a random CCA for the remaining students.
    for s in studentlist:
        rand=random.randrange(len(list(quotadict.keys())))
        if quotadict[list(quotadict.keys())[rand]] == 0:
            quotadict[list(quotadict.keys())[rand]] = quotadict[list(quotadict.keys())[rand]] + 1
            allocate(s, list(quotadict.keys())[rand], None)
            quotadict[list(quotadict.keys())[rand]] = quotadict[list(quotadict.keys())[rand]] - 1
# The quota becomes -1 such that each CCA does not get more than one extra student.

def stats_counter():
# This function counts choice and shortlist stats.
    sl_assign = len(slassignlist)
    choice_count = 0
    print("Number of non-DSA people who got in ccas which shortlisted them: " + str(sl_assign))
    for i in range(2,2+student_number):
        if classlist["H"+str(i)].value == 1:
            choice_count = choice_count + 1
    print("Percentage of people who got top choice: " + str(int(choice_count/choice_number * 100)) + "%")
    for i in range(2,2+student_number):
        if classlist["H"+str(i)].value == 2:
            choice_count = choice_count + 1
    print("Percentage of people who got top two choices: " + str(int(choice_count/choice_number * 100)) + "%")
    for i in range(2,2+student_number):
        if classlist["H"+str(i)].value == 3:
            choice_count = choice_count + 1
    print("Percentage of people who got top three choices: " + str(int(choice_count/choice_number * 100)) + "%")

allocate_dsa()
allocate_mep()
allocate_restshortlist()
allocate_restnormal()
allocate_restrandom()
stats_counter()
print(quotadict)
wb.save("allocated.xlsx")
