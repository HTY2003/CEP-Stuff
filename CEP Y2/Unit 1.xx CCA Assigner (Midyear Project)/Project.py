from openpyxl import *
import random

noofstudents = 0
noofchoices = 0

wb = load_workbook('cleaned2.xlsx')
classlist = wb['classlist']
choiceslist = wb['choices']
MEP = wb["MEP"]
DSA = wb["DSA"]

musiclist = ["SE","RV","GE","RIMB","RICO","RP"]         #for MEP

ccaswithshortlist = list(wb.get_sheet_names())          #to check if the cca has a shortlist for adding students(e.g swimming has no shortlist and hence no spreadsheet)
if "CO" in ccaswithshortlist:
    ccaswithshortlist.remove("CO")
    ccaswithshortlist.append("RICO")

slassignlist = []       
sllist = []

studentlist=[]
for i in range(2,2+500):
    if classlist["C"+str(i)].value != None:
        studentlist.append(classlist["C"+str(i)].value)
        noofstudents = noofstudents+1
for i in range(2,2+500):
    if choiceslist["B"+str(i)].value != None:
        noofchoices = noofchoices+1

namechangedict = {}                                     #made to account for the various inconsistences in name between spreadsheets
namechangedict["01SCOUT"] = "O1"
namechangedict["02SCOUT"] = "O2"

quotadict={}                                            #keeps track of all current CCA Quotas
quota=wb["ccaquota"]
for i in range(1,50):
    if isinstance(quota["E"+str(i)].value, int) == True:
        if quota["B"+str(i)].value == "RD":
                quotadict["Debater"]= quota["E"+str(i)].value
        else:
            quotadict[quota["B"+str(i)].value]=quota["E"+str(i)].value
                        
def allocate(name,CCA,choice,rank = None):
    if name in studentlist:
        studentlist.remove(name)
        for i in range(2,2+noofstudents):
            if classlist['C'+str(i)].value == name:
                classlist['G'+str(i)].value = CCA                         #updates the info for the student in the classlist
                classlist['H'+str(i)].value = choice
                classlist['I'+str(i)].value = rank
                if rank != None:
                    slassignlist.append(name)                             #adds everyone who got allocated by shortlist
                    if CCA == "RICO":
                        shortlist=wb["CO"]
                    else:
                        shortlist = wb[CCA]
                    for i in range(2,40):
                        if shortlist["B"+str(i)].value == name:
                            shortlist["D"+str(i)].value = "Allocated"     #makes it simpler to see who got allocated by shortlist for each CCA
        quotadict[CCA]=quotadict[CCA] - 1                                 #updates the dictionary and ccaquota spreadsheet
        for i in range(1,50):
            if CCA == "Debater":
                if quota["B"+str(i)].value == "RD":
                    quota["D"+str(i)].value = quota["D"+str(i)].value + 1
            elif quota["B"+str(i)].value == CCA:
                quota["D"+str(i)].value = quota["D"+str(i)].value + 1
        print(name,len(slassignlist))
def DSAAllo():                      #allocates DSA students
    print("DSA")
    for i in range(2,80):
        name = DSA["A"+str(i)].value
        CCA = DSA["C"+str(i)].value
        allocate(name,CCA,"DSA")

def MEPShortlist():                 #allocates MEP students based on whether 1)their first choice is music-based and 2)if they are in the shortlist for the cca
    print("MEP Shortlist")
    for i in range(2,40):
        name = MEP["A"+str(i)].value
        for n in range(2,2+noofchoices):
            if name == choiceslist["B"+str(n)].value:
                CCA = choiceslist["F"+str(n)].value
                if CCA in musiclist:
                    for r in range(2,40):
                        if CCA == "RICO":
                            shortlist = wb["CO"]
                        else:
                            shortlist = wb[CCA]
                        if shortlist["B"+str(i)].value == name:
                            if quotadict[CCA] > 0:
                                allocate(name,CCA,1,rank=r-1)

                                         
def RestShortlist():                #scans through the top rank in every cca and if the person has put it as their first choice then allocate
    print("Rest Shortlist")         #then it goes through the second rank in every cca and so on
    for c in range(6,9):            #then it does the same thing again except that it's the second choice(same thing with the third choice)
        for r in range(2,40):
            for CCA in list(quotadict.keys()):
                if CCA in ccaswithshortlist:
                    for n in range(2,2+noofchoices):
                        if CCA in list(namechangedict.keys()):
                            shortlist=wb[namechangedict[CCA]]
                            name = shortlist['B'+str(r)].value
                            if name == choiceslist["B"+str(n)].value:
                                if choiceslist.cell(row=n,column=c).value == namechangedict[CCA]:
                                    if quotadict[namechangedict[CCA]] > 0:
                                        allocate(name,namechangedict[CCA],c-5,rank=r-1)
                        else:
                            if CCA == "RICO":
                                shortlist=wb["CO"]
                            else:
                                shortlist=wb[CCA]
                            name = shortlist['B'+str(r)].value
                            if name == choiceslist["B"+str(n)].value:
                                if choiceslist.cell(row=n,column=c).value == CCA:
                                    if quotadict[CCA] > 0:
                                        allocate(name,CCA,c-5,rank=r-1)

def RestNormal():           #if slots are open in a choice, they get allocated
    print("Rest Normal")
    for c in range(6,15):   #if not, next choice
        for s in studentlist:
            for n in range(2,2+noofchoices):
                if s == choiceslist["B"+str(n)].value:
                    CCA = choiceslist.cell(row=n,column=c).value
                    if CCA in list(namechangedict.keys()):
                        if quotadict[namechangedict[CCA]] > 0:
                            allocate(s,namechangedict[CCA],c-5)
                    elif quotadict[CCA] > 0:
                            allocate(s,CCA,c-5)


def RestRandom():                                           #for those who did not hand in a choice or got really unlucky and all their choices' quotas were filled
    print("Rest Random")
    for CCA in list(quotadict.keys()):
        if quotadict[CCA] > 0:
            for repeat in range(quotadict[CCA]):            #allocates all remaining slots
                allocate(studentlist[0],CCA,None)
    for s in studentlist:
        rand=random.randrange(len(list(quotadict.keys())))  #adds some slots randomly if every CCA is full
        quotadict[list(quotadict.keys())[rand]] = quotadict[list(quotadict.keys())[rand]] + 1
        allocate(s,list(quotadict.keys())[rand],None)

def ShortlistCounter():                                 #doesn't count overlaps so if a student is shortlisted by 10 ccas, it only counts as 1
    for CCA in list(quotadict.keys()):
        if CCA in ccaswithshortlist:
            for x in range(2,40):
                if CCA == "RICO":
                    shortlist=wb["CO"]
                else:
                    shortlist=wb[CCA]
                if shortlist["B"+str(x)].value != None:
                    if shortlist["B"+str(x)].value not in sllist:
                        sllist.append(shortlist["B"+str(x)].value)

def StatsCounter():
    sl=len(sllist)
    slassign=len(slassignlist)
    choicecount=0
    print("Percentage of people who got in ccas which shortlisted them["+str(slassign)+"]/Total people in cca shortlists["+str(sl)+"]: " + str(int(slassign/sl * 100)) + "%")
    for i in range(2,2+noofstudents):
        if classlist["H"+str(i)].value == 1:
            choicecount = choicecount + 1
    print("Percentage of people who got top choice: " + str(int(choicecount/noofchoices * 100)) + "%")
    for i in range(2,2+noofstudents):
        if classlist["H"+str(i)].value == 2:
            choicecount = choicecount + 1
    print("Percentage of people who got top two choices: " + str(int(choicecount/noofchoices * 100)) + "%")
    for i in range(2,2+noofstudents):
        if classlist["H"+str(i)].value == 3:
            choicecount = choicecount + 1
    print("Percentage of people who got top three choices: " + str(int(choicecount/noofchoices * 100)) + "%")
    
DSAAllo()
MEPShortlist()
RestShortlist()
RestNormal()
RestRandom()
ShortlistCounter()
StatsCounter()

wb.save('done.xlsx')
